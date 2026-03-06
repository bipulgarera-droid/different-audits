import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def generate_audit_excel(audit_data):
    """
    Generates an Excel file from audit data.
    Returns a BytesIO object containing the Excel file.
    """
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo-600
    centered = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Helper to style headers
    def style_header(ws, row_idx=1):
        for cell in ws[row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border

    # =========================================================================
    # SHEET 1: SUMMARY
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Audit Summary"
    
    # Title
    ws_summary['A1'] = "SEO Audit Report"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Client Info (if available in future, passed in audit_data)
    
    # Scores (Placeholder or real if available)
    results = audit_data.get('results', {})
    summary = results.get('summary', {})
    categorized = results.get('categorized', {})
    
    ws_summary['A4'] = "Overall Health"
    ws_summary['B4'] = f"{summary.get('domain_rank', 'N/A')}" # Placeholder for score
    
    # Category Breakdown
    ws_summary['A6'] = "Category Breakdown"
    ws_summary['A6'].font = Font(bold=True)
    
    headers = ["Category", "Metric", "Issues", "Score"]
    ws_summary.append([]) # Spacer
    ws_summary.append(headers)
    style_header(ws_summary, 8)
    
    row = 9
    
    # Iterate through categories
    for category_name, metrics in categorized.items():
        for metric_key, data in metrics.items():
            ws_summary.cell(row=row, column=1, value=category_name.title())
            ws_summary.cell(row=row, column=2, value=metric_key.replace('_', ' ').title())
            ws_summary.cell(row=row, column=3, value=data.get('issues', 0))
            ws_summary.cell(row=row, column=4, value=data.get('score', '-'))
            row += 1

    # Adjust widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15

    # =========================================================================
    # SHEET 2: ISSUE DETAILS
    # =========================================================================
    ws_issues = wb.create_sheet("Issue Details")
    ws_issues.append(["Category", "Issue Type", "Priority", "URL", "Score/Details"])
    style_header(ws_issues, 1)
    
    for category_name, metrics in categorized.items():
        for metric_key, data in metrics.items():
            items = data.get('items', [])
            priority = data.get('priority', 'Medium') # Defaults might need populating in utils.py or here
            
            # If items is a list of strings (URLs)
            for item in items:
                if isinstance(item, str):
                    url = item
                    details = ""
                elif isinstance(item, dict):
                    url = item.get('url', 'N/A')
                    details = str(item)
                else:
                    url = str(item)
                    details = ""
                    
                ws_issues.append([
                    category_name.title(),
                    metric_key.replace('_', ' ').title(),
                    priority, # Note: Priority might not be deeply nested in 'categorized' in current utils.py structure, might need lookup
                    url,
                    details
                ])

    ws_issues.column_dimensions['A'].width = 15
    ws_issues.column_dimensions['B'].width = 30
    ws_issues.column_dimensions['C'].width = 15
    ws_issues.column_dimensions['D'].width = 60
    ws_issues.column_dimensions['E'].width = 30

    # =========================================================================
    # SHEET 3: ALL PAGES
    # =========================================================================
    ws_pages = wb.create_sheet("All Pages")
    ws_pages.append(["URL", "Status Code", "Title", "Meta Description", "H1", "Word Count", "Load Time (ms)"])
    style_header(ws_pages, 1)
    
    pages = results.get('pages', [])
    for page in pages:
        url = page.get('url', '')
        status = page.get('status_code', '-')
        
        # Meta might be nested or direct
        meta = page.get('meta', {})
        if not meta:
             # Fallback for dataforseo structure if different
             meta = page.get('checks', {}) # Sometimes checks contain meta info? No, usually 'meta' key.
        
        title = meta.get('title', '')
        desc = meta.get('description', '')
        h1 = meta.get('h1', '')
        word_count = meta.get('word_count', 0)
        
        # Performance
        speed = page.get('speed_metric', page.get('load_time', 0))
        
        ws_pages.append([url, status, title, desc, h1, word_count, speed])

    ws_pages.column_dimensions['A'].width = 60
    ws_pages.column_dimensions['B'].width = 10
    ws_pages.column_dimensions['C'].width = 40
    ws_pages.column_dimensions['D'].width = 40
    ws_pages.column_dimensions['E'].width = 30
    ws_pages.column_dimensions['F'].width = 15
    ws_pages.column_dimensions['G'].width = 15

    # Setup stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
